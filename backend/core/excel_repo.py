from __future__ import annotations

import os
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from . import solar_core as sc


def default_excel_path() -> Path:
    """
    Caminho padrão:
      Dimensionar_web/data/Planilha_de_Dimensionamento_Fortlev_Solar_20.xlsx
    Permite override com variável de ambiente EXCEL_PATH.
    """
    env = os.getenv("EXCEL_PATH")
    if env:
        return Path(env).expanduser().resolve()
    base_dir = Path(__file__).resolve().parents[2]  # .../Dimensionar_web
    return (base_dir / "data" / "Planilha_de_Dimensionamento_Fortlev_Solar_20.xlsx").resolve()


# lock simples para evitar duas escritas simultâneas no Excel
_excel_write_lock = threading.Lock()


def _encontrar_cabecalho(df_raw: pd.DataFrame, colunas_busca: List[str], min_match=3) -> Optional[int]:
    for i, row in df_raw.iterrows():
        vals = set(str(v).strip() for v in row.values if pd.notna(v) and str(v).strip() != "")
        hit = sum(1 for c in colunas_busca if c in vals)
        if hit >= min_match:
            return int(i)
    return None


def _limpar_df(df: pd.DataFrame, col_modelo: str) -> pd.DataFrame:
    df = df.loc[:, df.columns.notna()]
    df = df.loc[:, ~df.columns.astype(str).str.upper().str.startswith("UNNAMED")]
    df = df.loc[:, ~df.columns.astype(str).str.upper().str.startswith("_COL")]

    if col_modelo not in df.columns:
        raise RuntimeError(f"Coluna '{col_modelo}' não encontrada. Colunas: {list(df.columns)}")

    df = df[df[col_modelo].notna()].copy()
    df = df[df[col_modelo].astype(str).str.strip() != ""].copy()
    df = df[df[col_modelo].astype(str).str.upper().str.strip() != "MODELO"].copy()
    df.reset_index(drop=True, inplace=True)
    return df


def carregar_dados(excel_path: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not excel_path.exists():
        raise FileNotFoundError(f"Arquivo Excel não encontrado: {excel_path}")

    # INVERSORES (cabeçalho pode não estar na linha 1)
    df_raw_inv = pd.read_excel(excel_path, sheet_name="INVERSORES", header=None, dtype=str)
    idx = _encontrar_cabecalho(
        df_raw_inv,
        colunas_busca=[sc.CI_MODELO, sc.CI_FABRIC, sc.CI_VMAX, sc.CI_VMPP_MAX, sc.CI_VMIN_PART, sc.CI_POT_SAI],
        min_match=3,
    )
    if idx is None:
        raise RuntimeError("Não encontrei o cabeçalho na aba INVERSORES.")

    header = [
        str(v).strip() if pd.notna(v) and str(v).strip() != "" else f"_col{j}"
        for j, v in enumerate(df_raw_inv.iloc[idx])
    ]
    df_inv = df_raw_inv.iloc[idx + 1:].copy()
    df_inv.columns = header
    df_inv.reset_index(drop=True, inplace=True)
    df_inv = _limpar_df(df_inv, sc.CI_MODELO)
    if sc.CI_FABRIC in df_inv.columns:
        df_inv = df_inv[df_inv[sc.CI_FABRIC].notna()].copy()
        df_inv = df_inv[df_inv[sc.CI_FABRIC].astype(str).str.strip() != ""].copy()
        df_inv.reset_index(drop=True, inplace=True)

    # MODULOS (cabeçalho na primeira linha)
    df_mod = pd.read_excel(excel_path, sheet_name="MODULOS", header=0, dtype=str)
    df_mod.columns = [str(c).strip() for c in df_mod.columns]
    df_mod = _limpar_df(df_mod, sc.CM_MODELO)
    if sc.CM_FABRIC in df_mod.columns:
        df_mod = df_mod[df_mod[sc.CM_FABRIC].notna()].copy()
        df_mod = df_mod[df_mod[sc.CM_FABRIC].astype(str).str.strip() != ""].copy()
        df_mod.reset_index(drop=True, inplace=True)

    return df_inv, df_mod


# =========================
# Escrita no Excel
# =========================

def _mapear_colunas(ws: Worksheet, header_row: int) -> Dict[str, int]:
    colmap = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(header_row, col).value
        if v is None:
            continue
        name = str(v).strip()
        if name != "":
            colmap[name] = col
    return colmap


def _achar_linha_cabecalho_openpyxl(ws: Worksheet, colunas_busca: List[str], min_match=3, max_scan=60) -> Optional[int]:
    for r in range(1, min(ws.max_row, max_scan) + 1):
        vals = set()
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                vals.add(s)
        hit = sum(1 for col in colunas_busca if col in vals)
        if hit >= min_match:
            return r
    return None


def _proxima_linha_vazia(ws: Worksheet, col_modelo: int, start_row: int) -> int:
    r = start_row
    while True:
        v = ws.cell(r, col_modelo).value
        if v is None or str(v).strip() == "":
            return r
        r += 1


def append_produto_excel(excel_path: Path, tipo: str, dados: Dict[str, Any]) -> None:
    with _excel_write_lock:
        wb = load_workbook(excel_path)

        if tipo == "MODULO":
            if "MODULOS" not in wb.sheetnames:
                raise RuntimeError("Aba MODULOS não encontrada.")
            ws = wb["MODULOS"]
            header_row = 1
            colmap = _mapear_colunas(ws, header_row)
            if sc.CM_MODELO not in colmap:
                raise RuntimeError("Cabeçalho da aba MODULOS não contém a coluna MODELO.")
            row_out = _proxima_linha_vazia(ws, colmap[sc.CM_MODELO], header_row + 1)
            for k, v in dados.items():
                if k in colmap:
                    ws.cell(row_out, colmap[k]).value = v

        elif tipo == "INVERSOR":
            if "INVERSORES" not in wb.sheetnames:
                raise RuntimeError("Aba INVERSORES não encontrada.")
            ws = wb["INVERSORES"]
            header_row = _achar_linha_cabecalho_openpyxl(
                ws,
                colunas_busca=[sc.CI_MODELO, sc.CI_FABRIC, sc.CI_VMAX, sc.CI_VMPP_MAX, sc.CI_POT_SAI],
                min_match=3,
            )
            if header_row is None:
                raise RuntimeError("Não consegui localizar o cabeçalho na aba INVERSORES.")
            colmap = _mapear_colunas(ws, header_row)
            if sc.CI_MODELO not in colmap:
                raise RuntimeError("Cabeçalho da aba INVERSORES não contém a coluna MODELO.")
            row_out = _proxima_linha_vazia(ws, colmap[sc.CI_MODELO], header_row + 1)
            for k, v in dados.items():
                if k in colmap:
                    ws.cell(row_out, colmap[k]).value = v

        else:
            raise ValueError("tipo inválido. Use 'MODULO' ou 'INVERSOR'.")

        wb.save(excel_path)


# =========================
# Store (cache)
# =========================

@dataclass
class DataStore:
    excel_path: Path = default_excel_path()
    df_inv: Optional[pd.DataFrame] = None
    df_mod: Optional[pd.DataFrame] = None
    last_error: Optional[str] = None

    def load(self) -> None:
        try:
            self.df_inv, self.df_mod = carregar_dados(self.excel_path)
            self.last_error = None
        except Exception as e:
            self.df_inv, self.df_mod = None, None
            self.last_error = str(e)

    def ensure_loaded(self) -> None:
        if self.df_inv is None or self.df_mod is None:
            self.load()