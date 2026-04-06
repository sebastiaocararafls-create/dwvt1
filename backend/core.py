from __future__ import annotations

import math
import threading
from pathlib import Path
from typing import Optional, Dict, Any, List, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

# ============================================================
# CONFIG
# ============================================================

BASE_DIR = Path(__file__).resolve().parents[1]
ARQUIVO_EXCEL = BASE_DIR / "data" / "Planilha_de_Dimensionamento_Fortlev_Solar_20.xlsx"
TEMP_STC = 25.0

_excel_write_lock = threading.Lock()

# INVERSORES (nomes conforme Excel)
CI_MODELO = "MODELO"
CI_FABRIC = "FABRICANTE"
CI_POT_ENT = "MAX. POT. ENTRADA (KWP)"
CI_N_ENT = "Nº ENTRADAS INVERSOR"
CI_N_MPPT = "Nº MPPTS"
CI_POT_SAI = "MAX. POT. SAÍDA (KW)"
CI_VAC = "TENSÃO SAÍDA (V)"
CI_VMAX = "TENSÃO MÁX. ENTRADA (V)"
CI_VMPP_MAX = "TENSÃO MÁX. MPP (V)"
CI_VMIN_PART = "TENSÃO MÍN PARTIDA. (V)"
CI_IMAX = "CORRENTE MÁX. ENTRADA FV (A)"
CI_ICC = "CORRENTE MÁX. CURTO (A)"
CI_CFG_MPPT = "CONFIGURAÇÃO MPPTs"

# MODULOS
CM_MODELO = "MODELO"
CM_FABRIC = "FABRICANTE"
CM_POT = "POTÊNCIA (KWP)"  # na planilha: Wp
CM_TIPO = "TIPO CÉLULA"
CM_VOC = "TENSÃO ABERTO (Voc)"
CM_VMP = "TENSÃO OP. STC (Vmp)"
CM_IMP = "CORRENTE (A)"
CM_ISC = "CORRENTE CURTO (A)"
CM_EFIC = "EFICIÊNCIA (%)"
CM_COEF_ISC = "COEF. TEMP. CORR. CURTO (%/°C)"
CM_COEF_VOC = "COEF. TEMP. TENSÃO ABERTO (%/°C)"
CM_BIFACIAL = "MONO/BIFACIAL"

# ============================================================
# UTIL
# ============================================================

def to_float(v, default=0.0) -> float:
    try:
        if v is None:
            return float(default)
        if isinstance(v, (int, float)):
            f = float(v)
            return float(default) if math.isnan(f) else f
        s = str(v).strip()
        if s == "" or s.lower() == "nan":
            return float(default)
        s = s.replace(",", ".")
        for sep in ("/", " "):
            if sep in s:
                s = s.split(sep)[0].strip()
                break
        return float(s)
    except Exception:
        return float(default)

def to_int(v, default=0) -> int:
    try:
        return int(float(str(v).replace(",", ".")))
    except Exception:
        return int(default)

def parse_config_mppt(config_str: str) -> List[int]:
    if not config_str:
        return []
    s = str(config_str).strip()
    if s == "" or s.lower() == "nan":
        return []
    out = []
    for part in s.split("/"):
        part = part.strip()
        if not part:
            continue
        if " " in part:
            part = part.split(" ")[0].strip()
        part = part.replace(",", ".")
        try:
            n = int(float(part))
            if n > 0:
                out.append(n)
        except Exception:
            pass
    return out

def entradas_total_inversor(inv: Dict[str, Any]) -> int:
    cfg = str(inv.get(CI_CFG_MPPT, "")).strip()
    parsed = parse_config_mppt(cfg)
    if parsed:
        return max(1, sum(parsed))
    n = to_int(inv.get(CI_N_ENT, 0), 0)
    return n if n > 0 else 1

def overload_max(inv: Dict[str, Any]) -> float:
    pin = to_float(inv.get(CI_POT_ENT, 0), 0.0)
    pout = to_float(inv.get(CI_POT_SAI, 0), 0.0)
    if pin <= 0 or pout <= 0:
        return 1.8
    return max(pin / pout, 1.0)

# ============================================================
# EXCEL LOAD
# ============================================================

def _encontrar_cabecalho(df_raw: pd.DataFrame, colunas_busca: List[str], min_match=3) -> Optional[int]:
    for i, row in df_raw.iterrows():
        vals = set(str(v).strip() for v in row.values if pd.notna(v) and str(v).strip() != "")
        hit = sum(1 for c in colunas_busca if c in vals)
        if hit >= min_match:
            return int(i)
    return None

def _limpar_df(df: pd.DataFrame, col_modelo: str) -> pd.DataFrame:
    df = df.loc[:, df.columns.notna()]
    df = df.loc[:, ~df.columns.astype(str).str.upper().str.startswith("UNNAMED")]
    df = df.loc[:, ~df.columns.astype(str).str.upper().str.startswith("_COL")]

    if col_modelo not in df.columns:
        raise RuntimeError(f"Coluna '{col_modelo}' não encontrada. Colunas: {list(df.columns)}")

    df = df[df[col_modelo].notna()].copy()
    df = df[df[col_modelo].astype(str).str.strip() != ""].copy()
    df = df[df[col_modelo].astype(str).str.upper().str.strip() != "MODELO"].copy()
    df.reset_index(drop=True, inplace=True)
    return df

def carregar_dados() -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not ARQUIVO_EXCEL.exists():
        raise FileNotFoundError(f"Arquivo Excel não encontrado: {ARQUIVO_EXCEL}")

    df_raw_inv = pd.read_excel(ARQUIVO_EXCEL, sheet_name="INVERSORES", header=None, dtype=str)
    idx = _encontrar_cabecalho(
        df_raw_inv,
        colunas_busca=[CI_MODELO, CI_FABRIC, CI_VMAX, CI_VMPP_MAX, CI_VMIN_PART, CI_POT_SAI],
        min_match=3
    )
    if idx is None:
        raise RuntimeError("Não encontrei o cabeçalho na aba INVERSORES.")

    header = [
        str(v).strip() if pd.notna(v) and str(v).strip() != "" else f"_col{j}"
        for j, v in enumerate(df_raw_inv.iloc[idx])
    ]
    df_inv = df_raw_inv.iloc[idx + 1:].copy()
    df_inv.columns = header
    df_inv.reset_index(drop=True, inplace=True)
    df_inv = _limpar_df(df_inv, CI_MODELO)
    if CI_FABRIC in df_inv.columns:
        df_inv = df_inv[df_inv[CI_FABRIC].notna()].copy()
        df_inv = df_inv[df_inv[CI_FABRIC].astype(str).str.strip() != ""].copy()
        df_inv.reset_index(drop=True, inplace=True)

    df_mod = pd.read_excel(ARQUIVO_EXCEL, sheet_name="MODULOS", header=0, dtype=str)
    df_mod.columns = [str(c).strip() for c in df_mod.columns]
    df_mod = _limpar_df(df_mod, CM_MODELO)
    if CM_FABRIC in df_mod.columns:
        df_mod = df_mod[df_mod[CM_FABRIC].notna()].copy()
        df_mod = df_mod[df_mod[CM_FABRIC].astype(str).str.strip() != ""].copy()
        df_mod.reset_index(drop=True, inplace=True)

    return df_inv, df_mod

# ============================================================
# EXCEL WRITE
# ============================================================

def _mapear_colunas(ws: Worksheet, header_row: int) -> Dict[str, int]:
    colmap = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(header_row, col).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            colmap[name] = col
    return colmap

def _achar_linha_cabecalho_openpyxl(ws: Worksheet, colunas_busca: List[str], min_match=3, max_scan=60) -> Optional[int]:
    for r in range(1, min(ws.max_row, max_scan) + 1):
        vals = set()
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                vals.add(s)
        hit = sum(1 for col in colunas_busca if col in vals)
        if hit >= min_match:
            return r
    return None

def _proxima_linha_vazia(ws: Worksheet, col_modelo: int, start_row: int) -> int:
    r = start_row
    while True:
        v = ws.cell(r, col_modelo).value
        if v is None or str(v).strip() == "":
            return r
        r += 1

def append_produto_excel(tipo: str, dados: Dict[str, Any]) -> None:
    with _excel_write_lock:
        wb = load_workbook(ARQUIVO_EXCEL)

        if tipo == "MODULO":
            ws = wb["MODULOS"]
            header_row = 1
            colmap = _mapear_colunas(ws, header_row)
            row_out = _proxima_linha_vazia(ws, colmap[CM_MODELO], header_row + 1)
            for k, v in dados.items():
                if k in colmap:
                    ws.cell(row_out, colmap[k]).value = v

        elif tipo == "INVERSOR":
            ws = wb["INVERSORES"]
            header_row = _achar_linha_cabecalho_openpyxl(
                ws,
                colunas_busca=[CI_MODELO, CI_FABRIC, CI_VMAX, CI_VMPP_MAX, CI_POT_SAI],
                min_match=3
            )
            if header_row is None:
                raise RuntimeError("Não consegui localizar o cabeçalho na aba INVERSORES.")
            colmap = _mapear_colunas(ws, header_row)
            row_out = _proxima_linha_vazia(ws, colmap[CI_MODELO], header_row + 1)
            for k, v in dados.items():
                if k in colmap:
                    ws.cell(row_out, colmap[k]).value = v
        else:
            raise ValueError("tipo inválido. Use 'MODULO' ou 'INVERSOR'.")

        wb.save(ARQUIVO_EXCEL)